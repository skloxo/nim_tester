# -*- coding: utf-8 -*-
"""
Web 服务器：FastAPI + SSE 实时进度推送
提供 Web UI 界面，自动驱动全流程
"""
import asyncio
import json
import logging
import sys
import webbrowser
from pathlib import Path
from typing import Any, Dict, List

import uvicorn
import yaml
from fastapi import FastAPI, Body
from pydantic import BaseModel
from typing import Optional
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from tester.network import NetworkSelector
from tester.model_fetcher import ModelFetcher
from tester.categorizer import ModelCategorizer
from tester.runner import TestRunner
from tester.scorer import score_model, rank_category
from tester.use_case import infer_use_cases
from tester.db import init_db, save_run, finish_run, save_model_result, save_model_results_batch, list_runs, get_run_results
from tester.meta_fetcher import bulk_init_meta

logger = logging.getLogger("web_server")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# 初始化数据库
init_db()

PROFILES_PATH = Path("profiles.json")

app = FastAPI(title="API Model Tester")


class RunConfig(BaseModel):
    """UI 上可覆盖的运行配置"""
    api_keys: Optional[list] = None
    base_url: Optional[str] = None
    proxy: Optional[str] = None
    concurrency: Optional[int] = None
    required_only: Optional[bool] = None
    rate_limit_per_key: Optional[int] = None   # 每 key 每分钟请求上限，0=不限

# ── 全局状态 ─────────────────────────────────────────────────────────────────
import collections as _collections
_events = _collections.deque(maxlen=2000)  # 上限防止长时间运行内存膨胀
_run_state = {"running": False, "results": None, "scored": None, "run_id": None}


def _emit(event: dict):
    """追加事件到历史列表"""
    _events.append(event)


def _make_log_emit():
    """创建一个向 SSE 流写日志的 logging Handler"""
    class SseHandler(logging.Handler):
        def emit(self, record):
            _emit({"type": "log", "level": record.levelname.lower(),
                   "msg": self.format(record)})
    h = SseHandler()
    h.setFormatter(logging.Formatter("%(asctime)s %(message)s", datefmt="%H:%M:%S"))
    return h


# ── SSE 端点 ─────────────────────────────────────────────────────────────────
@app.get("/api/events")
async def sse_events(since: int = 0):
    async def generate():
        idx = since
        while True:
            events_list = list(_events)   # deque 转 list，支持下标访问
            if idx < len(events_list):
                yield f"data: {json.dumps(events_list[idx], ensure_ascii=False)}\n\n"
                idx += 1
            else:
                if events_list and events_list[-1].get("type") == "complete":
                    break
                await asyncio.sleep(0.15)
    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── 启动测试 ─────────────────────────────────────────────────────────────────
@app.post("/api/run")
async def start_run(body: RunConfig = Body(default=RunConfig())):
    if _run_state["running"]:
        return JSONResponse({"ok": False, "msg": "测试已在运行中"})
    _events.clear()
    _run_state["running"] = True
    _run_state["results"] = None
    _run_state["scored"] = None
    asyncio.create_task(_run_pipeline(body.model_dump(exclude_none=True)))
    return JSONResponse({"ok": True})


async def _run_pipeline(overrides: dict = None):
    config = yaml.safe_load(Path("config.yaml").read_text(encoding="utf-8"))
    if overrides:
        # UI 层面可覆盖的配置项
        if overrides.get("api_keys"):
            config["api_keys"] = [k.strip() for k in overrides["api_keys"] if k.strip()]
        if overrides.get("base_url"):
            config["api"]["base_url"] = overrides["base_url"].rstrip("/")
        if overrides.get("proxy") is not None:
            config["network"]["proxy"] = overrides["proxy"]
        if overrides.get("concurrency"):
            config["testing"]["concurrency"] = int(overrides["concurrency"])
        if overrides.get("required_only") is not None:
            config["testing"]["required_only"] = overrides["required_only"] in (True, "true", 1)
        if overrides.get("rate_limit_per_key") is not None:
            config["testing"]["rate_limit_per_key"] = int(overrides["rate_limit_per_key"])

    # 注册 SSE log handler
    sse_h = _make_log_emit()
    root_log = logging.getLogger()
    root_log.addHandler(sse_h)

    def progress(event: dict):
        _emit(event)

    try:
        from datetime import datetime
        import uuid
        run_id = uuid.uuid4().hex[:12]
        run_started = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        profile_name = overrides.get("_profile", "") if overrides else ""
        save_run(run_id, profile_name, config["api"]["base_url"], config, run_started)
        _run_state["run_id"] = run_id
        total_models = 0  # 提前初始化，防止 Step4 前异常导致 finish_run NameError

        # Step 1: 测速
        _emit({"type": "step", "step": 1, "label": "自动测速..."})
        selector = NetworkSelector(config)
        best_mode, latency = await selector.select_best()
        _emit({"type": "step_done", "step": 1,
               "label": f"最优路径: {best_mode}（{latency:.0f}ms）"})

        # Step 2: 拉取模型
        _emit({"type": "step", "step": 2, "label": "拉取全量模型列表..."})
        fetcher = ModelFetcher(config, best_mode)
        models = await fetcher.fetch_all()
        model_ids = [m["id"] for m in models]
        _emit({"type": "step_done", "step": 2, "label": f"获取 {len(models)} 个模型"})

        # Step 2.5: 元数据初始化（首次全量 + 增量补漏）
        _emit({"type": "step", "step": 25, "label": f"同步模型元数据（{len(models)} 个）..."})
        proxy = config.get("network", {}).get("proxy", "")
        api_key = config["api_keys"][0] if config.get("api_keys") else ""
        meta_done = [0]

        async def _meta_progress(done, total, mid):
            meta_done[0] = done
            if done % 10 == 0 or done == total:
                _emit({"type": "meta_progress", "done": done, "total": total})

        all_meta = await bulk_init_meta(
            model_ids, proxy=proxy,
            base_url=config["api"]["base_url"], api_key=api_key,
            progress_cb=_meta_progress,
        )
        _run_state["meta"] = all_meta
        _emit({"type": "step_done", "step": 25,
               "label": f"元数据就绪（{len(all_meta)} 个模型）"})

        # Step 3: 分组
        _emit({"type": "step", "step": 3, "label": "模型智能分组..."})
        categorizer = ModelCategorizer(config)
        groups = categorizer.categorize(models)
        cat_conf = config["model_categories"]
        groups_info = {
            k: {"count": len(v), "desc": cat_conf.get(k, {}).get("description", k)}
            for k, v in groups.items()
        }
        _emit({"type": "groups", "groups": groups_info})
        _emit({"type": "step_done", "step": 3, "label": f"共 {len(groups)} 个分类"})

        # Step 4: 测试
        _emit({"type": "step", "step": 4, "label": "分类测试中..."})
        total_models = sum(len(v) for v in groups.values())
        done_count = [0]

        async def on_model_done(result_list):
            done_count[0] += 1
            if result_list:
                mid = result_list[0]["model_id"]
                cat = result_list[0]["category"]
                passed = sum(1 for r in result_list if r.get("status") == "pass")
                _emit({"type": "model_done", "model_id": mid, "category": cat,
                       "passed": passed, "total": len(result_list),
                       "done": done_count[0], "total_models": total_models})

        runner = TestRunner(config, best_mode, progress_callback=on_model_done)
        results = await runner.run_all(groups)
        _emit({"type": "step_done", "step": 4, "label": "测试完成"})

        # Step 5: 评分 + 报告
        _emit({"type": "step", "step": 5, "label": "生成评分与报告..."})
        scored = _build_scored(results, config)
        _run_state["results"] = results
        _run_state["scored"] = scored

        # 写入 SQLite 历史（O-5: 单一批量事务）
        try:
            batch_items = [
                (mid, cat, m, m.get("results", []))
                for cat, data in scored.items()
                for mid, m in data["models"].items()
            ]
            save_model_results_batch(run_id, batch_items)
            finish_run(run_id, datetime.now().strftime("%Y-%m-%dT%H:%M:%S"), total_models)
        except Exception as db_err:
            logger.warning(f"DB 持久化失败（不影响主流程）: {db_err}")

        # 保存 JSON
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = Path("results") / ts
        out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / "results.json").write_text(
            json.dumps(results, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        (out_dir / "scored.json").write_text(
            json.dumps(scored, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )

        # 保存 Excel
        try:
            _save_excel(results, scored, config, out_dir)
        except Exception as e:
            logger.warning(f"Excel 生成失败: {e}")

        _emit({"type": "step_done", "step": 5, "label": f"报告已保存: {out_dir}"})
        scored_summary = {
            cat: {
                "description": data["description"],
                "models": {
                    mid: {k: v for k, v in m.items() if k != "results"}
                    for mid, m in data["models"].items()
                }
            }
            for cat, data in scored.items()
        }
        _emit({"type": "complete", "scored": scored_summary,
               "out_dir": str(out_dir), "run_id": run_id})

    except Exception as e:
        logger.exception(f"流程异常: {e}")
        _emit({"type": "error", "msg": str(e)})
    finally:
        root_log.removeHandler(sse_h)
        _run_state["running"] = False


def _build_scored(results: dict, config: dict) -> dict:
    """对每个分类的模型计算评分、排名、使用场景"""
    cat_conf = config["model_categories"]
    output = {}
    for category, cat_results in results.items():
        desc = cat_conf.get(category, {}).get("description", category)
        # 按模型 ID 聚合
        by_model: Dict[str, list] = {}
        for r in cat_results:
            by_model.setdefault(r["model_id"], []).append(r)

        # 评分
        model_scores = {mid: score_model(rs) for mid, rs in by_model.items()}
        ranked = rank_category(model_scores)

        # 使用场景
        for mid, rs in by_model.items():
            use_case_str = infer_use_cases(rs, category)
            # 存为 list 便于 DB 序列化和前端处理；"⚠️" 开头的不可用标记单独一项
            ranked[mid]["use_cases"] = (
                use_case_str.split("、") if use_case_str and "⚠️" not in use_case_str
                else ([use_case_str] if use_case_str else [])
            )
            ranked[mid]["use_cases_str"] = use_case_str   # 保留字符串供 Excel 使用
            ranked[mid]["results"] = rs

        output[category] = {"description": desc, "models": ranked}
    return output


def _save_excel(results, scored, config, out_dir: Path):
    import re
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    cat_conf = config["model_categories"]

    # 总览 sheet
    ws = wb.active
    ws.title = "总览"
    headers = ["分类", "模型数", "平均分", "最高分", "第一名", "通过率"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for cat, data in scored.items():
        models = data["models"]
        scores = [m["score"] for m in models.values()]
        top = min(models.values(), key=lambda m: m["rank"])
        top_name = min(models, key=lambda mid: models[mid]["rank"])
        cat_results = results.get(cat, [])
        total = len(cat_results)
        passed = sum(1 for r in cat_results if r.get("status") == "pass")
        ws.append([
            data["description"], len(models),
            round(sum(scores) / len(scores), 1) if scores else 0,
            max(scores) if scores else 0,
            top_name,
            f"{passed/total*100:.1f}%" if total else "N/A",
        ])

    # 每个分类一个 sheet
    for cat, data in scored.items():
        safe = re.sub(r'[/\\?*\[\]:]', '_', data["description"])[:31]
        ws2 = wb.create_sheet(title=safe)
        h2 = ["排名", "模型 ID", "评分", "等级", "通过/总用例", "TPS", "推荐场景"]
        ws2.append(h2)
        _style_header(ws2, 1, len(h2))
        for mid, m in sorted(data["models"].items(), key=lambda x: x[1]["rank"]):
            rank_str = {1: "🥇", 2: "🥈", 3: "🥉"}.get(m["rank"], f"#{m['rank']}")
            # use_cases_str 是字符串，use_cases 是 list；Excel 显示字符串
            uc_display = m.get("use_cases_str") or "、".join(m.get("use_cases", []))
            ws2.append([
                rank_str, mid, m["score"], m["grade"],
                f"{m['passed']}/{m['total']}", m["avg_tps"], uc_display,
            ])
            row = ws2.max_row
            grade_colors = {"S": "7B2FBE", "A": "1D4ED8", "B": "047857",
                            "C": "B45309", "D": "9CA3AF", "F": "DC2626"}
            color = grade_colors.get(m["grade"], "9CA3AF")
            ws2.cell(row, 3).fill = PatternFill("solid", fgColor=color)
            ws2.cell(row, 3).font = Font(color="FFFFFF", bold=True)

        for col in range(1, len(h2) + 1):
            ws2.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = 20

    wb.save(out_dir / "report.xlsx")
    logger.info(f"Excel 报告: {out_dir / 'report.xlsx'}")


def _style_header(ws, row, col_count):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1E40AF")
    for c in range(1, col_count + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


# ── 查询接口 ─────────────────────────────────────────────────────────────────
@app.get("/api/results")
async def get_results():
    if not _run_state["scored"]:
        return JSONResponse({"ok": False, "msg": "暂无结果"})
    return JSONResponse({"ok": True, "scored": _run_state["scored"]})


@app.get("/api/config")
async def get_config():
    cfg = yaml.safe_load(Path("config.yaml").read_text(encoding="utf-8"))
    return JSONResponse(cfg)


@app.get("/api/export/excel")
async def export_excel():
    """下载最新的 Excel 报告（从 results/ 目录）"""
    from fastapi.responses import FileResponse
    reports = sorted(Path("results").glob("*/report.xlsx"))
    if not reports:
        return JSONResponse({"ok": False, "msg": "暂无报告，请先运行测试"}, status_code=404)
    latest = reports[-1]
    return FileResponse(str(latest), filename=f"model_report_{latest.parent.name}.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/export/excel/{run_id}")
async def export_excel_run(run_id: str):
    """按 run_id 从数据库重新生成并下载 Excel 报告（含元数据）"""
    import tempfile, openpyxl, sqlite3 as _sq
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import FileResponse

    try:
        results = get_run_results(run_id)
        if not results:
            return JSONResponse({"ok": False, "msg": "未找到该 Run 的测试数据"}, status_code=404)

        # ── 加载分类中文描述 ──────────────────────────────────────────────
        from tester.db import DB_PATH
        cat_descs: dict = {}
        try:
            cfg = _load_config()
            for k, v in cfg.get("model_categories", {}).items():
                cat_descs[k] = v.get("description", k)
        except Exception:
            pass

        # ── 从 model_meta_cache 批量读取元数据 ───────────────────────────
        meta_map: dict = {}
        try:
            con = _sq.connect(str(DB_PATH)); con.row_factory = _sq.Row
            mids = list({r["model_id"] for r in results})
            placeholders = ",".join("?" * len(mids))
            mrows = con.execute(
                f"SELECT model_id, param_count, max_context, release_date, hf_pipeline_tag "
                f"FROM model_meta_cache WHERE model_id IN ({placeholders})", mids
            ).fetchall()
            con.close()
            for mr in mrows:
                meta_map[mr["model_id"]] = dict(mr)
        except Exception as e:
            logger.warning(f"元数据读取失败: {e}")

        # ── 按分类分组并排序 ──────────────────────────────────────────────
        by_cat: dict = {}
        for row in results:
            by_cat.setdefault(row["category"], []).append(row)

        # ── 创建 Excel ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "模型测试报告"

        HEADERS = [
            "模型 ID", "分类", "排名", "评分", "等级",
            "通过用例", "总用例数", "通过率(%)",
            "Tokens/s (均值)", "平均响应时间(ms)",
            "参数量(B)", "上下文长度(tokens)", "发布日期", "推荐场景"
        ]
        # 表头样式
        hdr_fill = PatternFill("solid", fgColor="1E2A40")
        hdr_font = Font(bold=True, color="A5B4FC")
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        row_idx = 2
        for cat, rows in by_cat.items():
            cat_cn = cat_descs.get(cat, cat)
            for row in sorted(rows, key=lambda r: r.get("rank") or 99):
                mid = row["model_id"]
                meta = meta_map.get(mid, {})
                uc = row.get("use_cases") or []
                uc_str = "、".join(uc) if isinstance(uc, list) else str(uc)

                # 从 results_json 计算 avg_elapsed
                avg_elapsed = None
                try:
                    raw = json.loads(row.get("results_json") or "[]")
                    elapsed_vals = [r.get("elapsed_ms") for r in raw
                                    if r.get("elapsed_ms") and r.get("status") == "pass"]
                    if elapsed_vals:
                        avg_elapsed = round(sum(elapsed_vals) / len(elapsed_vals), 0)
                except Exception:
                    pass

                passed = row.get("passed", 0)
                total = row.get("total", 0)
                pass_rate = round(passed / total * 100, 1) if total else 0
                tps = row.get("avg_tps") or None

                ws.append([
                    mid,
                    cat_cn,
                    row.get("rank"),
                    row.get("score"),
                    row.get("grade"),
                    passed,
                    total,
                    pass_rate,
                    round(tps, 2) if tps else None,
                    avg_elapsed,
                    meta.get("param_count"),
                    meta.get("max_context"),
                    meta.get("release_date"),
                    uc_str,
                ])
                row_idx += 1

        # 自动列宽
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        # 写临时文件
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name); tmp.close()
        return FileResponse(tmp.name, filename=f"model_report_{run_id}.xlsx",
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        logger.exception(f"Excel 生成失败: {e}")
        return JSONResponse({"ok": False, "msg": str(e)}, status_code=500)



# ── 历史记录接口 ──────────────────────────────────────────────────────────────
@app.get("/api/history")
async def get_history():
    """返回最近 20 次测试摘要"""
    try:
        runs = list_runs(limit=20)
        return JSONResponse({"ok": True, "runs": runs})
    except Exception as e:
        return JSONResponse({"ok": False, "msg": str(e)})


@app.get("/api/history/{run_id}")
async def get_history_detail(run_id: str):
    """返回某次测试的详细模型结果"""
    try:
        results = get_run_results(run_id)
        return JSONResponse({"ok": True, "run_id": run_id, "results": results})
    except Exception as e:
        return JSONResponse({"ok": False, "msg": str(e)})


# ── API 档案管理接口 ───────────────────────────────────────────────────────────
def _load_profiles() -> dict:
    if PROFILES_PATH.exists():
        try:
            return json.loads(PROFILES_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"profiles": [], "last_used": ""}


def _save_profiles(data: dict):
    PROFILES_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


@app.get("/api/profiles")
async def get_profiles():
    return JSONResponse(_load_profiles())


class ProfileModel(BaseModel):
    name: str
    base_url: str
    api_keys: list
    proxy: Optional[str] = ""
    rate_limit_per_key: Optional[int] = 40


@app.post("/api/profiles")
async def save_profile(p: ProfileModel):
    """新建或更新档案（按 name 覆盖）"""
    data = _load_profiles()
    profiles = data.get("profiles", [])
    # 找到同名档案更新，否则追加
    for i, existing in enumerate(profiles):
        if existing["name"] == p.name:
            profiles[i] = p.model_dump()
            data["last_used"] = p.name
            _save_profiles(data)
            return JSONResponse({"ok": True, "action": "updated"})
    profiles.append(p.model_dump())
    data["last_used"] = p.name
    _save_profiles(data)
    return JSONResponse({"ok": True, "action": "created"})


@app.delete("/api/profiles/{name}")
async def delete_profile(name: str):
    data = _load_profiles()
    before = len(data.get("profiles", []))
    data["profiles"] = [p for p in data.get("profiles", []) if p["name"] != name]
    if data.get("last_used") == name:
        data["last_used"] = data["profiles"][0]["name"] if data["profiles"] else ""
    _save_profiles(data)
    deleted = before - len(data["profiles"])
    return JSONResponse({"ok": True, "deleted": deleted})


# ── 模型目录信息 ──────────────────────────────────────────────────────────────
@app.get("/api/catalog/stats")
async def catalog_stats():
    """返回本地 model_catalog.json 统计信息"""
    cat_path = Path("data/model_catalog.json")
    if not cat_path.exists():
        return JSONResponse({"ok": False, "msg": "catalog 未生成，请运行 tools/build_catalog.py"})
    try:
        catalog = json.loads(cat_path.read_text(encoding="utf-8"))
        has_params = sum(1 for e in catalog.values() if e.get("param_count"))
        has_date = sum(1 for e in catalog.values() if e.get("release_date"))
        has_hf = sum(1 for e in catalog.values() if e.get("hf_pipeline_tag"))
        updated = max((e.get("updated_at", "") for e in catalog.values()), default="")
        return JSONResponse({
            "ok": True,
            "total": len(catalog),
            "has_params": has_params,
            "has_date": has_date,
            "hf_matched": has_hf,
            "updated_at": updated,
        })
    except Exception as e:
        return JSONResponse({"ok": False, "msg": str(e)})


@app.get("/")
async def index():
    return HTMLResponse(Path("static/index.html").read_text(encoding="utf-8"))


# ── 启动入口 ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import threading, time
    def _open_browser():
        time.sleep(1.5)
        webbrowser.open("http://localhost:28080")
    threading.Thread(target=_open_browser, daemon=True).start()
    uvicorn.run(app, host="0.0.0.0", port=28080, log_level="warning")
