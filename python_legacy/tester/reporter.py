# -*- coding: utf-8 -*-
"""报告生成模块：输出 JSON + Excel + 控制台摘要"""

import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

logger = logging.getLogger(__name__)


class Reporter:
    """生成测试报告（JSON / Excel / 控制台）"""

    def __init__(self, config: dict):
        self.config = config
        self.output_dir = Path(config["testing"]["output_dir"])
        self.save_json = config["testing"].get("save_json", True)
        self.save_excel = config["testing"].get("save_excel", True)

    def save(self, results: Dict[str, Any]) -> Path:
        """保存所有报告，返回输出目录"""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_dir = self.output_dir / ts
        run_dir.mkdir(parents=True, exist_ok=True)

        if self.save_json:
            self._save_json(results, run_dir)

        if self.save_excel:
            try:
                self._save_excel(results, run_dir)
            except ImportError:
                logger.warning("⚠️  未安装 openpyxl，跳过 Excel 报告（pip install openpyxl）")

        return run_dir

    def _save_json(self, results: dict, run_dir: Path):
        path = run_dir / "results.json"
        with open(path, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2, default=str)
        logger.info(f"  📄 JSON 报告: {path}")

    def _save_excel(self, results: dict, run_dir: Path):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        # ── 总览 Sheet ────────────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "总览"
        summary_headers = ["分类", "描述", "模型数", "用例总数", "通过", "失败", "跳过", "通过率"]
        ws_summary.append(summary_headers)
        self._style_header(ws_summary, 1, len(summary_headers))

        cat_conf = self.config["model_categories"]

        for category, cat_results in results.items():
            desc = cat_conf.get(category, {}).get("description", category)
            model_ids = list({r["model_id"] for r in cat_results})
            total = len(cat_results)
            passed = sum(1 for r in cat_results if r.get("status") == "pass")
            failed = sum(1 for r in cat_results if r.get("status") in ("fail", "error"))
            skipped = sum(1 for r in cat_results if r.get("status") == "skip")
            rate = f"{passed/total*100:.1f}%" if total > 0 else "N/A"
            ws_summary.append([category, desc, len(model_ids), total, passed, failed, skipped, rate])

        # ── 明细 Sheet（每个分类一个 Sheet）─────────────────────────────
        detail_headers = [
            "模型 ID", "测试用例", "状态", "耗时(ms)", "Token/s",
            "输出预览", "finish_reason", "工具调用", "has_reasoning",
            "has_chinese", "supports_vlm", "dimension", "备注",
        ]
        for category, cat_results in results.items():
            desc = cat_conf.get(category, {}).get("description", category)
            desc = cat_conf.get(category, {}).get("description", category)
            # Excel Sheet 标题不允许含 / \ ? * [ ] : 等字符，做清理
            import re
            safe_title = re.sub(r'[/\\?*\[\]:]', '_', desc)[:31]
            ws = wb.create_sheet(title=safe_title)
            ws.append(detail_headers)
            self._style_header(ws, 1, len(detail_headers))

            for r in cat_results:
                row = [
                    r.get("model_id", ""),
                    r.get("case_name", ""),
                    r.get("status", ""),
                    r.get("elapsed_ms", ""),
                    r.get("tps", ""),
                    r.get("content_preview", "")[:100] if r.get("content_preview") else "",
                    r.get("finish_reason", ""),
                    str(bool(r.get("tool_calls"))),
                    str(r.get("has_reasoning", "")),
                    str(r.get("has_chinese", "")),
                    str(r.get("supports_vlm", "")),
                    r.get("dimension", ""),
                    r.get("reason", ""),
                ]
                ws.append(row)

                # 着色
                last_row = ws.max_row
                status = r.get("status", "")
                fill = None
                if status == "pass":
                    fill = PatternFill("solid", fgColor="C6EFCE")
                elif status in ("fail", "error"):
                    fill = PatternFill("solid", fgColor="FFC7CE")
                elif status == "skip":
                    fill = PatternFill("solid", fgColor="FFEB9C")
                if fill:
                    for col in range(1, len(detail_headers) + 1):
                        ws.cell(last_row, col).fill = fill

            # 自适应列宽
            for col_idx in range(1, len(detail_headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

        path = run_dir / "report.xlsx"
        wb.save(path)
        logger.info(f"  📊 Excel 报告: {path}")

    @staticmethod
    def _style_header(ws, row: int, col_count: int):
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="4472C4")
        for col in range(1, col_count + 1):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def print_summary(self, results: Dict[str, Any]):
        """在控制台打印简洁的测试摘要"""
        cat_conf = self.config["model_categories"]
        print("\n" + "═" * 70)
        print(f"{'分类':<20} {'模型数':>6} {'用例数':>7} {'通过':>6} {'失败':>6} {'通过率':>8}")
        print("─" * 70)

        for category, cat_results in results.items():
            desc = cat_conf.get(category, {}).get("description", category)
            model_ids = list({r["model_id"] for r in cat_results})
            total = len(cat_results)
            passed = sum(1 for r in cat_results if r.get("status") == "pass")
            failed = total - passed
            rate = f"{passed/total*100:.1f}%" if total > 0 else "N/A"
            print(f"{desc:<20} {len(model_ids):>6} {total:>7} {passed:>6} {failed:>6} {rate:>8}")

        print("═" * 70)

        # 打印每个分类中的失败模型
        print("\n🔴 失败/错误详情：")
        any_fail = False
        for category, cat_results in results.items():
            fails = [r for r in cat_results if r.get("status") in ("fail", "error")]
            for f in fails:
                any_fail = True
                print(f"  ❌ [{category}] {f['model_id']} → {f['case_name']}: {f.get('reason', '')}")
        if not any_fail:
            print("  ✅ 全部通过！")
